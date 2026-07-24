"""加速器学会（PASJ）のプログラムを構造化 JSON に変換する。"""

import argparse
import csv
import json
import re
from datetime import date, datetime
from pathlib import Path


# 日本語→英語キーのマッピング（ここにない通常列は出力しない）
key_map = {
    "ログインID": "login_id",
    "演題登録番号": "submission_id",
    "参加登録番号": "registration_id",
    "講演番号": "talk_id",
    "セッション": "session",
    "確定発表形式": "confirmed_format",
    "開催日": "date",
    "発表時間": "time",
    "会場": "room",
    "姓": "last_name",
    "名": "first_name",
    "姓(かな)": "last_name_kana",
    "名(かな)": "first_name_kana",
    "姓(英語)": "last_name_en",
    "名(英語)": "first_name_en",
    "名(ミドルネーム)": "middle_name",
    "所属機関名": "affiliation",
    "部署名・学部名": "department",
    "発表形式": "presentation_type",
    "題目": "title_ja",
    "題目（english）": "title_en",
    "要旨": "abstract_ja",
    "発表希望カテゴリー（第１希望）": "category_1",
    "発表希望カテゴリー（第２希望）": "category_2",
    "年会賞への応募応募": "award_entry",
    "賞状の敬称": "honorific",
    "通信欄": "note",
    "著者情報": "author_text_html",
    "著者情報（英語）": "author_text_en_html",
    "著者所属情報": "affiliation_text_html",
    "著者所属情報（英語）": "affiliation_text_en_html",
    "登録日時": "created_at",
    "更新日時": "updated_at",
}

# 英語キーの連名者フィールド（ここにない連名者列は出力しない）
coauthor_fields = {
    "区分": "role",
    "姓": "last_name",
    "名": "first_name",
    "姓カナ": "last_name_kana",
    "名カナ": "first_name_kana",
    "姓英語": "last_name_en",
    "名英語": "first_name_en",
    "ミドルネーム": "middle_name",
    "所属": "affiliation",
}

COAUTHOR_PATTERN = re.compile(r"連名者\[(\d+)\](.+)")


def cell_to_string(value):
    """CSV を介した場合と同様に、セル値を文字列へそろえる。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def format_xlsx_value(header, value, workbook_epoch):
    """xlsx の開催日を「8月26日」形式へ変換する。"""
    if header != "開催日" or value in (None, ""):
        return value

    if isinstance(value, (int, float)):
        from openpyxl.utils.datetime import from_excel

        value = from_excel(value, workbook_epoch)

    if isinstance(value, (date, datetime)):
        return f"{value.month}月{value.day}日"

    return value


def convert_row(row):
    converted = {}
    coauthors = []

    for key, value in row.items():
        if not key:
            continue

        value = cell_to_string(value)
        match = COAUTHOR_PATTERN.fullmatch(key)
        if match:
            index = int(match.group(1))
            field_en = coauthor_fields.get(match.group(2))
            if field_en:
                while len(coauthors) < index:
                    coauthors.append({"index": len(coauthors) + 1})
                coauthors[index - 1][field_en] = value
        elif key in key_map:
            # 未定義の列を無視することで、従来の「余計な列を削除」を行う。
            converted[key_map[key]] = value

    coauthors = [
        coauthor
        for coauthor in coauthors
        if any(value.strip() for key, value in coauthor.items() if key != "index")
    ]
    converted["coauthors"] = coauthors
    return converted


def read_csv(file_path):
    with open(file_path, mode="r", encoding="utf-8-sig", newline="") as csv_file:
        yield from csv.DictReader(csv_file)


def read_xlsx(file_path, sheet_name=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "xlsx の読み込みには openpyxl が必要です。"
            " 'python -m pip install openpyxl' でインストールしてください。"
        ) from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name is not None and sheet_name not in workbook.sheetnames:
            names = ", ".join(workbook.sheetnames)
            raise ValueError(f"シート '{sheet_name}' がありません。利用可能なシート: {names}")

        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return

        # 同名列が複数ある場合は、Excel 上で左側の列を採用する。
        unique_columns = []
        seen_headers = set()
        for column_index, header in enumerate(headers):
            if header is None:
                continue
            header = str(header).strip()
            if header and header not in seen_headers:
                seen_headers.add(header)
                unique_columns.append((column_index, header))

        for values in rows:
            if not any(value is not None for value in values):
                continue
            yield {
                header: format_xlsx_value(
                    header,
                    values[column_index] if column_index < len(values) else None,
                    workbook.epoch,
                )
                for column_index, header in unique_columns
            }
    finally:
        workbook.close()


def convert_to_json(input_file_path, json_file_path, sheet_name=None):
    suffix = Path(input_file_path).suffix.lower()
    if suffix == ".csv":
        if sheet_name is not None:
            raise ValueError("--sheet は xlsx 入力でのみ指定できます。")
        rows = read_csv(input_file_path)
    elif suffix == ".xlsx":
        rows = read_xlsx(input_file_path, sheet_name)
    else:
        raise ValueError("入力ファイルは .csv または .xlsx を指定してください。")

    converted_data = [convert_row(row) for row in rows]
    with open(json_file_path, mode="w", encoding="utf-8") as json_file:
        json.dump(converted_data, json_file, indent=4, ensure_ascii=False)


def parse_args():
    parser = argparse.ArgumentParser(
        description="加速器学会のプログラム（CSV/xlsx）を構造化 JSON に変換します。"
    )
    parser.add_argument("input", help="入力する .csv または .xlsx ファイル")
    parser.add_argument("output", help="出力する .json ファイル")
    parser.add_argument(
        "--sheet",
        help="xlsx から読み込むシート名（省略時は先頭シート）",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        convert_to_json(args.input, args.output, args.sheet)
    except (OSError, RuntimeError, ValueError) as exc:
        raise SystemExit(f"エラー: {exc}") from exc
